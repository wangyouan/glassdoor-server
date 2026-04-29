from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
import pyarrow.parquet as pq
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

MAJOR_PATH = Path(
    "/data/disk4/workspace/projects/glassdoor/outputs/firm_year_glassdoor_major_customer.parquet"
)
UNION_PATH = Path(
    "/data/disk4/workspace/projects/glassdoor/outputs/firm_year_glassdoor_union.parquet"
)
AGG_SCRIPT_PATH = Path(
    "/data/disk4/workspace/projects/glassdoor/src/build_firm_year_aggregates.py"
)
OUTPUT_XLSX_PATH = Path(
    "/data/disk4/workspace/projects/glassdoor/outputs/firm_year_glassdoor_variable_definitions.xlsx"
)

JOB_GROUPS = [
    "sales",
    "rd",
    "operations",
    "admin",
    "management",
    "senior",
    "entry_level",
    "client_facing",
    "high_hc",
    "specialized",
    "core_business",
    "performance_linked",
    "compliance",
    "low_level_core",
    "high_level_core",
    "non_core_staff",
]

RATING_LABELS = {
    "GD_rating": "overall rating",
    "GD_outlook": "business outlook",
    "GD_career_opp": "career opportunities rating",
    "GD_ceo": "CEO rating",
    "GD_comp_benefit": "compensation and benefits rating",
    "GD_culture": "culture and values rating",
    "GD_diversity": "diversity and inclusion rating",
    "GD_recommend": "recommend-to-friend rating",
    "GD_senior_mgmt": "senior management rating",
    "GD_wlb": "work-life balance rating",
}

SUBGROUP_RATING_VARS = [
    "GD_rating",
    "GD_career_opp",
    "GD_comp_benefit",
    "GD_senior_mgmt",
    "GD_wlb",
    "GD_culture",
    "GD_diversity",
]

TEXT_LABELS = {
    "avg_summary_len": "Mean review-summary character length",
    "avg_advice_len": "Mean review-advice character length",
    "avg_pros_len": "Mean review-pros character length",
    "avg_cons_len": "Mean review-cons character length",
    "avg_total_text_len": "Mean total review text character length",
    "avg_pros_word_count": "Mean review-pros word count",
    "avg_cons_word_count": "Mean review-cons word count",
}


def read_parquet_schema(path: Path) -> Tuple[List[str], Dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Parquet file not found: {path}")
    pf = pq.ParquetFile(path)
    names = pf.schema.names
    types = {field.name: str(field.type) for field in pf.schema_arrow}
    return names, types


def build_variable_row(var: str, dtype_note: str) -> Dict[str, str]:
    row = {
        "variable": var,
        "category": "Other",
        "definition": "Variable from firm-year Glassdoor aggregation output.",
        "construction": "Constructed in build_firm_year_aggregates.py.",
        "source_review_level_variables": "",
        "aggregation_level": "gvkey x review_year",
        "missing_value_rule": "See construction rule in aggregation script.",
        "notes": dtype_note,
    }

    if var == "gvkey":
        row.update(
            {
                "category": "Identifier",
                "definition": "Compustat firm identifier.",
                "construction": "Passed through from review-level gvkey and grouped at firm-year.",
                "source_review_level_variables": "gvkey",
                "missing_value_rule": "Rows with missing gvkey are dropped before aggregation.",
            }
        )
        return row

    if var == "review_year":
        row.update(
            {
                "category": "Identifier",
                "definition": "Calendar year of Glassdoor review.",
                "construction": "Passed through from review-level review_year and grouped at firm-year.",
                "source_review_level_variables": "review_year",
                "missing_value_rule": "Rows with missing review_year are dropped before aggregation.",
            }
        )
        return row

    fixed = {
        "n_reviews": (
            "Review count",
            "Number of Glassdoor reviews for a firm-year.",
            "Count of review rows within each gvkey x review_year group.",
            "all reviews",
            "Always non-missing for retained firm-years.",
        ),
        "n_current_emp": (
            "Employment composition",
            "Number of reviews by current employees.",
            "Sum of indicator is_current_employee == 1 within firm-year.",
            "is_current_employee",
            "0 when no current-employee reviews are observed.",
        ),
        "n_former_emp": (
            "Employment composition",
            "Number of reviews by former employees.",
            "Sum of indicator is_former_employee == 1 within firm-year.",
            "is_former_employee",
            "0 when no former-employee reviews are observed.",
        ),
        "pct_current": (
            "Employment composition",
            "Share of current-employee reviews.",
            "n_current_emp / (n_current_emp + n_former_emp) when denominator > 0.",
            "is_current_employee, is_former_employee",
            "Missing when n_current_emp + n_former_emp == 0.",
        ),
        "n_unique_rcid": (
            "Identifier",
            "Number of distinct rcid values in a firm-year.",
            "Distinct count of review-level rcid within firm-year.",
            "rcid",
            "0 if all rcid values are missing.",
        ),
        "n_unique_company_names": (
            "Company name",
            "Number of distinct company names in a firm-year.",
            "Distinct count of non-missing review-level company strings.",
            "company",
            "0 if all company values are missing.",
        ),
        "company_name_mode": (
            "Company name",
            "Most frequent company name observed in a firm-year.",
            "Mode of non-missing review-level company; ties broken lexicographically.",
            "company",
            "Missing when no non-missing company names are available.",
        ),
        "ultimate_parent_company_name_mode": (
            "Company name",
            "Most frequent ultimate parent company name observed in a firm-year.",
            "Mode of non-missing review-level ultimate_parent_company_name; ties broken lexicographically.",
            "ultimate_parent_company_name",
            "Missing when no non-missing ultimate parent names are available.",
        ),
        "has_10_reviews": (
            "Review-count flag",
            "Indicator for firm-years with at least 10 reviews.",
            "1 if n_reviews >= 10 else 0.",
            "n_reviews",
            "Always non-missing (0/1).",
        ),
        "has_25_reviews": (
            "Review-count flag",
            "Indicator for firm-years with at least 25 reviews.",
            "1 if n_reviews >= 25 else 0.",
            "n_reviews",
            "Always non-missing (0/1).",
        ),
        "has_50_reviews": (
            "Review-count flag",
            "Indicator for firm-years with at least 50 reviews.",
            "1 if n_reviews >= 50 else 0.",
            "n_reviews",
            "Always non-missing (0/1).",
        ),
    }

    if var in fixed:
        cat, definition, construction, source, miss = fixed[var]
        row.update(
            {
                "category": cat,
                "definition": definition,
                "construction": construction,
                "source_review_level_variables": source,
                "missing_value_rule": miss,
            }
        )
        return row

    if var in RATING_LABELS:
        src = var.replace("GD_", "gd_")
        row.update(
            {
                "category": "Rating mean",
                "definition": f"Firm-year mean {RATING_LABELS[var]}.",
                "construction": f"Mean of review-level {src} within gvkey x review_year.",
                "source_review_level_variables": src,
                "missing_value_rule": f"Missing when no non-missing {src} in the firm-year.",
            }
        )
        return row

    m_count = re.fullmatch(r"n_(GD_[A-Za-z0-9_]+)", var)
    if m_count:
        gd = m_count.group(1)
        src = gd.replace("GD_", "gd_")
        row.update(
            {
                "category": "Rating count",
                "definition": f"Count of non-missing review-level {src} used to construct {gd}.",
                "construction": f"Non-missing count of {src} within firm-year.",
                "source_review_level_variables": src,
                "missing_value_rule": "Always non-missing; 0 when no usable observations.",
            }
        )
        return row

    if var in TEXT_LABELS:
        src = var.replace("avg_", "")
        row.update(
            {
                "category": "Text length",
                "definition": TEXT_LABELS[var],
                "construction": f"Mean of review-level {src} within firm-year.",
                "source_review_level_variables": src,
                "missing_value_rule": f"Missing when no non-missing {src} in the firm-year.",
            }
        )
        return row

    m_pct = re.fullmatch(r"pct_([A-Za-z0-9_]+)", var)
    if m_pct and m_pct.group(1) in JOB_GROUPS:
        grp = m_pct.group(1)
        row.update(
            {
                "category": "Job composition",
                "definition": f"Share of reviews classified into {grp} job group.",
                "construction": f"Mean of review-level job_{grp} indicator within firm-year.",
                "source_review_level_variables": f"job_{grp}",
                "missing_value_rule": "Missing only if n_reviews is zero (normally non-missing).",
            }
        )
        return row

    m_n_grp = re.fullmatch(r"n_([A-Za-z0-9_]+)_reviews", var)
    if m_n_grp and m_n_grp.group(1) in JOB_GROUPS:
        grp = m_n_grp.group(1)
        row.update(
            {
                "category": "Subgroup review count",
                "definition": f"Number of reviews in the {grp} job group.",
                "construction": f"Count of reviews with job_{grp} == 1 within firm-year.",
                "source_review_level_variables": f"job_{grp}",
                "missing_value_rule": "Always non-missing; 0 when subgroup has no reviews.",
            }
        )
        return row

    m_sub = re.fullmatch(r"([A-Za-z0-9_]+)_(GD_[A-Za-z0-9_]+)", var)
    if m_sub and m_sub.group(1) in JOB_GROUPS and m_sub.group(2) in SUBGROUP_RATING_VARS:
        grp = m_sub.group(1)
        gd = m_sub.group(2)
        src = gd.replace("GD_", "gd_")
        row.update(
            {
                "category": "Subgroup rating",
                "definition": f"Mean {gd} among reviews in {grp} job group.",
                "construction": f"Mean of {src} among reviews with job_{grp} == 1 within firm-year.",
                "source_review_level_variables": f"{src}, job_{grp}",
                "missing_value_rule": "Missing when subgroup has no non-missing rating observations.",
            }
        )
        return row

    return row


def build_variable_definitions_df(
    columns: List[str],
    major_types: Dict[str, str],
    union_types: Dict[str, str],
) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for c in columns:
        dtype_major = major_types.get(c)
        dtype_union = union_types.get(c)

        if dtype_major and dtype_union:
            dtype_note = (
                f"dtype_major={dtype_major}; dtype_union={dtype_union}"
                if dtype_major != dtype_union
                else f"dtype={dtype_major}"
            )
        elif dtype_major:
            dtype_note = f"dtype_major={dtype_major}; missing_in_union"
        else:
            dtype_note = f"dtype_union={dtype_union}; missing_in_major"

        rows.append(build_variable_row(c, dtype_note))

    return pd.DataFrame(rows)


def build_aggregation_rules_df() -> pd.DataFrame:
    rules = [
        ("Unit of observation", "gvkey x review_year"),
        ("Rating mean variables", "Firm-year means of review-level gd_* ratings."),
        ("Rating count variables", "n_GD_* counts are non-missing review counts used in each mean."),
        ("Text variables", "avg_* variables are firm-year means of review-level text-length measures."),
        ("Job composition variables", "pct_* variables are firm-year shares/means of job-group dummies."),
        (
            "Subgroup rating variables",
            "<group>_GD_* variables are means among reviews where corresponding job dummy equals 1.",
        ),
        ("Subgroup count variables", "n_<group>_reviews are counts of reviews in each job group."),
        (
            "Review-count flags",
            "has_10_reviews / has_25_reviews / has_50_reviews are indicators based on n_reviews thresholds.",
        ),
        (
            "Representative company names",
            "Mode strings among non-missing names with lexicographic tie-break in aggregation script.",
        ),
        (
            "Source script",
            str(AGG_SCRIPT_PATH),
        ),
    ]
    return pd.DataFrame(rules, columns=["rule", "description"])


def build_job_groups_df() -> pd.DataFrame:
    rows = []
    for grp in JOB_GROUPS:
        subgroup_vars = [f"{grp}_{rv}" for rv in SUBGROUP_RATING_VARS]
        rows.append(
            {
                "group_name": grp,
                "dummy_variable": f"job_{grp}",
                "share_variable": f"pct_{grp}",
                "subgroup_review_count_variable": f"n_{grp}_reviews",
                "subgroup_rating_variables_created": ", ".join(subgroup_vars),
                "short_interpretation": f"Reviews classified into {grp} job group.",
            }
        )
    return pd.DataFrame(rows)


def apply_sheet_formatting(output_path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Column widths
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells[:2000]:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            width = min(max(12, max_len + 2), 70)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Wrap text for long-description columns
        header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        wrap_cols = [
            "definition",
            "construction",
            "source_review_level_variables",
            "missing_value_rule",
            "notes",
            "description",
            "subgroup_rating_variables_created",
            "short_interpretation",
        ]
        for name in wrap_cols:
            if name in header_map:
                cidx = header_map[name]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=cidx).alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )

    wb.save(output_path)


def main() -> None:
    major_cols, major_types = read_parquet_schema(MAJOR_PATH)
    union_cols, union_types = read_parquet_schema(UNION_PATH)

    # Keep major schema order first, then append columns only in union.
    ordered_cols = list(major_cols)
    union_only = [c for c in union_cols if c not in set(major_cols)]
    ordered_cols.extend(union_only)

    var_df = build_variable_definitions_df(ordered_cols, major_types, union_types)
    rules_df = build_aggregation_rules_df()
    job_df = build_job_groups_df()

    OUTPUT_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_XLSX_PATH, engine="openpyxl") as writer:
        var_df.to_excel(writer, sheet_name="Variable_Definitions", index=False)
        rules_df.to_excel(writer, sheet_name="Aggregation_Rules", index=False)
        job_df.to_excel(writer, sheet_name="Job_Groups", index=False)

    apply_sheet_formatting(OUTPUT_XLSX_PATH)

    print(f"Documented variables: {len(var_df):,}")
    print(f"Output file: {OUTPUT_XLSX_PATH}")


if __name__ == "__main__":
    main()
